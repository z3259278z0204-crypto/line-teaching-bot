import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def generate_excel(records, filepath):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '教具記錄'

    header_fill = PatternFill('solid', fgColor='4472C4')
    header_font = Font(color='FFFFFF', bold=True, size=12)
    center = Alignment(horizontal='center', vertical='center')
    left_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)
    alt_fill = PatternFill('solid', fgColor='DCE6F1')
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ['週次／日期', '教具名稱', '活動目標／備註', '記錄日期']
    widths = [18, 40, 48, 16]

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 28

    for i, r in enumerate(records, 2):
        date_str = r['created_at'][:10] if r.get('created_at') else ''
        row_data = [r['week'], r['tools'], r.get('notes', ''), date_str]
        fill = alt_fill if i % 2 == 0 else None
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.alignment = left_wrap
            cell.border = border
            if fill:
                cell.fill = fill
        ws.row_dimensions[i].height = 22

    total_row = len(records) + 3
    cell = ws.cell(row=total_row, column=1, value=f'共 {len(records)} 筆記錄')
    cell.font = Font(bold=True)

    wb.save(filepath)
